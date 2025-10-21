# apps/reports/generators/excel_generator.py
"""
SMGI Backend - Excel Report Generator
Sistema de Monitoreo Geoespacial Inteligente
Generador de informes en formato Excel (.xlsx)
"""
import logging
import io
import time
import os
from typing import Dict, Any, Optional, List, Union
from datetime import timedelta
from django.utils import timezone
from django.utils.translation import gettext_lazy as _

# Biblioteca para generar Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None

# from xlsxwriter import Workbook # Otra opción
# HAS_XLSXWRITER = True

from apps.reports.generators.base_generator import BaseReportGenerator
from apps.reports.models import Report, GeneratedReport, ReportFormat, ReportStatus


logger = logging.getLogger('apps.reports.generators.excel')


class ExcelReportGenerator(BaseReportGenerator):
    """
    Generador de informes en formato Excel (.xlsx) usando openpyxl.
    """

    def __init__(
        self,
        name: str = "Excel Report Generator",
        description: str = "Generates reports in Excel (.xlsx) format",
        is_active: bool = True,
        default_options: Optional[Dict[str, Any]] = None
    ):
        """
        Inicializa el generador de informes Excel.

        Args:
            name (str): Nombre del generador.
            description (str): Descripción del generador.
            is_active (bool): Indica si el generador está activo.
            default_options (Optional[Dict[str, Any]]): Opciones por defecto.
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required to use ExcelReportGenerator. Please install it: pip install openpyxl")
        
        super().__init__(
            name=name,
            description=description,
            format_type=ReportFormat.EXCEL,
            is_active=is_active,
            default_options=default_options or {
                'include_header': True,
                'auto_adjust_column_width': True,
                'freeze_header': True,
                'apply_styles': True,
                'include_summary_sheet': True,
                'include_charts': False, # Requiere datos específicos
                'include_images': False, # Requiere datos específicos
                'sheet_name_prefix': 'Data',
                'summary_sheet_name': 'Summary',
                'chart_sheet_name': 'Charts',
                'image_sheet_name': 'Images'
            }
        )
        self.logger = logging.getLogger(f'{self.__class__.__module__}.{self.__class__.__name__}')

    def generate(self, report: Report, data: Dict[str, Any], options: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Genera un informe en formato Excel (.xlsx).

        Args:
            report (Report): Instancia del modelo Report que se está generando.
            data (Dict[str, Any]): Datos para generar el informe.
                                   Debe contener 'sheets' como lista de diccionarios
                                   con 'name' y 'data' (lista de diccionarios).
                                   Ej: {'sheets': [{'name': 'Sheet1', 'data': [{'col1': 'val1', 'col2': 'val2'}, ...]}]}
            options (Optional[Dict[str, Any]]): Opciones específicas para esta generación.

        Returns:
            Dict[str, Any]: Resultado de la generación.
        """
        start_time = time.time()
        self.logger.info(f"Starting Excel report generation for {report.name}")
        
        # Merge options
        merged_options = self.default_options.copy()
        if options:
            merged_options.update(options)
        
        try:
            # Validate input data
            if not self.validate_input_data(data):
                return self.handle_error(ValueError("Invalid input data"), "Data validation failed")
            
            # Prepare output path
            output_path = self.prepare_output_path(report, suffix='.xlsx')
            
            # Create Excel workbook in memory
            wb = openpyxl.Workbook()
            ws_main = wb.active
            ws_main.title = merged_options.get('sheet_name_prefix', 'Data') + "_1"
            
            # Get sheets data
            sheets_data = data.get('sheets', [])
            if not sheets_data:
                # Fallback: create a sheet with report data if no sheets are provided
                sheets_data = [{'name': 'Report Data', 'data': data.get('report_data', [])}]
            
            total_records = 0
            sheet_count = 0
            
            # Create worksheets
            for i, sheet_info in enumerate(sheets_data):
                sheet_name = sheet_info.get('name', f'Sheet{i+1}')
                sheet_data = sheet_info.get('data', [])
                
                if i == 0:
                    ws = ws_main
                    ws.title = sheet_name[:31] # Excel sheet name max length is 31 chars
                else:
                    ws = wb.create_sheet(title=sheet_name[:31])
                
                if not sheet_data:
                    self.logger.warning(f"No data provided for sheet '{sheet_name}'. Creating empty sheet.")
                    continue
                
                # Write header row
                if merged_options.get('include_header', True) and sheet_data:
                    headers = list(sheet_data[0].keys())
                    ws.append(headers)
                    total_records += 1 # Header row counts as a record in Excel terms
                    
                    if merged_options.get('apply_styles', True):
                        self._apply_header_styles(ws, len(headers))
                
                # Write data rows
                for row_data in sheet_data:
                    row_values = [row_data.get(header, '') for header in headers] if merged_options.get('include_header', True) else list(row_data.values())
                    ws.append(row_values)
                    total_records += 1
                
                # Auto-adjust column widths
                if merged_options.get('auto_adjust_column_width', True):
                    self._auto_adjust_column_width(ws)
                
                # Freeze header row
                if merged_options.get('freeze_header', True) and merged_options.get('include_header', True):
                    ws.freeze_panes = 'A2'
                
                sheet_count += 1
            
            # Add summary sheet if requested
            if merged_options.get('include_summary_sheet', True):
                ws_summary = wb.create_sheet(title=merged_options.get('summary_sheet_name', 'Summary')[:31])
                ws_summary.append(['Summary'])
                ws_summary['A1'].font = Font(bold=True, size=14)
                ws_summary.append(['Total Sheets', sheet_count])
                ws_summary.append(['Total Records', total_records])
                ws_summary.append(['Generated At', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
                ws_summary.append(['Report Name', report.name])
                ws_summary.append(['Report ID', str(report.id)])
                
                # Auto-adjust column widths for summary
                if merged_options.get('auto_adjust_column_width', True):
                    self._auto_adjust_column_width(ws_summary)
            
            # Add charts sheet if requested and data is available
            # This is a placeholder. Real chart creation requires specific data structures.
            # if merged_options.get('include_charts', False):
            #     ws_charts = wb.create_sheet(title=merged_options.get('chart_sheet_name', 'Charts')[:31])
            #     # Chart creation logic would go here
            #     # Requires chart data in 'data' parameter
            
            # Add images sheet if requested and data is available
            # This is a placeholder. Real image insertion requires specific data structures.
            # if merged_options.get('include_images', False):
            #     ws_images = wb.create_sheet(title=merged_options.get('image_sheet_name', 'Images')[:31])
            #     # Image insertion logic would go here
            #     # Requires image data (paths, binary data) in 'data' parameter
            
            # Save workbook to file
            wb.save(output_path)
            
            # Get file size
            file_size_bytes = os.path.getsize(output_path)
            
            # Calculate duration
            duration_ms = int((time.time() - start_time) * 1000)
            
            self.logger.info(f"Excel report generated successfully for {report.name} in {duration_ms} ms")
            
            return {
                'success': True,
                'file_path': output_path,
                'file_size_bytes': file_size_bytes,
                'duration_ms': duration_ms,
                'records': total_records,
                'pages': sheet_count, # Using sheet count as pages
                'error': None
            }
            
        except Exception as e:
            error_msg = f"Error generating Excel report for {report.name}: {e}"
            self.logger.error(error_msg)
            return self.handle_error(e, "Excel generation failed")
        finally:
            # Cleanup temp files (handled by context manager or explicit call)
            # self.cleanup_temp_files() # No es necesario llamarlo aquí si se usa context manager
            pass

    def _apply_header_styles(self, worksheet, column_count: int):
        """
        Aplica estilos al encabezado de una hoja de trabajo.

        Args:
            worksheet: Hoja de trabajo de openpyxl.
            column_count (int): Número de columnas en el encabezado.
        """
        try:
            # Definir estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar estilos a la primera fila (encabezado)
            for col_num in range(1, column_count + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
        except Exception as e:
            self.logger.warning(f"Could not apply header styles: {e}")

    def _auto_adjust_column_width(self, worksheet):
        """
        Ajusta automáticamente el ancho de las columnas según el contenido.

        Args:
            worksheet: Hoja de trabajo de openpyxl.
        """
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50) # Limitar ancho máximo
                
        except Exception as e:
            self.logger.warning(f"Could not auto-adjust column width: {e}")

    def validate_input_data(self, data: Dict[str, Any]) -> bool:
        """
        Valida los datos de entrada específicos para generación de Excel.

        Args:
            data (Dict[str, Any]): Datos para validar.

        Returns:
            bool: True si los datos son válidos, False en caso contrario.
        """
        # Llamar al método de validación base
        if not super().validate_input_data(data):
            return False
        
        # Validación específica para Excel: verificar estructura de 'sheets'
        sheets_data = data.get('sheets')
        if sheets_data is not None:
            if not isinstance(sheets_data, list):
                self.logger.error("Excel generator expects 'sheets' to be a list.")
                return False
            
            for i, sheet_info in enumerate(sheets_data):
                if not isinstance(sheet_info, dict):
                    self.logger.error(f"Sheet info at index {i} must be a dictionary.")
                    return False
                
                sheet_name = sheet_info.get('name')
                sheet_data = sheet_info.get('data')
                
                if sheet_name is not None and not isinstance(sheet_name, str):
                    self.logger.error(f"Sheet name at index {i} must be a string.")
                    return False
                
                if sheet_data is not None and not isinstance(sheet_data, list):
                    self.logger.error(f"Sheet data at index {i} must be a list.")
                    return False
                
                # Validar que cada fila en sheet_data sea un dict
                if sheet_data:
                    for j, row in enumerate(sheet_data):
                        if not isinstance(row, dict):
                            self.logger.error(f"Row {j} in sheet {i} data must be a dictionary.")
                            return False
        
        return True
